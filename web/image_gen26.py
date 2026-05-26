"""
 — Image generation and diagram helpers.
 routes /api/generate-image, /api/image-providers, and /api/diagramit.
"""
import os
import base64
import sys
from pathlib import Path

# Load .env from project root (parent of web/)
sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv
load_dotenv()

import google.genai as genai
import google.genai.types as genai_types

GEMINI_API_KEY  = os.environ.get("GEMINI_API_KEY")
OPENAI_API_KEY  = os.environ.get("OPENAI_API_KEY")
AALTO_API_KEY   = os.environ.get("AALTO_OPENAI_API_KEY")  # kept for reference; Aalto deprecated DALL-E

_gemini_client = genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None


def generate_image_gemini(prompt: str) -> str:
    """Generate image with Gemini Flash image generation. Returns base64-encoded PNG bytes."""
    if not _gemini_client:
        raise ValueError("GEMINI_API_KEY not set")
    response = _gemini_client.models.generate_content(
        model="gemini-2.0-flash",
        contents=prompt,
        config=genai_types.GenerateContentConfig(
            response_modalities=["IMAGE", "TEXT"],
        ),
    )
    for part in response.candidates[0].content.parts:
        print("Part type:", type(part), "has inline_data:", part.inline_data is not None)


        if part.inline_data is not None:
            img_data = part.inline_data.data
            if isinstance(img_data, bytes):
                return base64.b64encode(img_data).decode("utf-8")
            return img_data  # already a base64 string
    raise ValueError("Gemini returned no image in the response")


def generate_image_openai(prompt: str) -> str:
    """Generate image with DALL-E 3 via direct OpenAI API. Returns base64-encoded PNG bytes."""
    if not OPENAI_API_KEY:
        raise ValueError("OPENAI_API_KEY not set")
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    response = client.images.generate(
        model="dall-e-3",
        prompt=prompt,
        size="1024x1024",
        quality="standard",
        n=1,
        response_format="b64_json",
    )
    return response.data[0].b64_json


def providers_available() -> dict:
    """Return which providers have API keys configured."""
    return {
        "gemini": bool(GEMINI_API_KEY),
        "openai": bool(OPENAI_API_KEY),
    }


def generate_diagram_excel(user_prompt: str):
    """Build diagram assumptions Excel from user prompt. Returns (BytesIO, model_used)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from model_client26 import generate_diagram_data

    diagram_data, model_used = generate_diagram_data(user_prompt)

    central_concept  = diagram_data.get('central_concept', 'Concept')
    meta_assumptions = diagram_data.get('meta_assumptions', [])[:3]
    subtopics        = diagram_data.get('subtopics', [])

    ROWS_PER_BLOCK = 3
    GAP = 1

    left_sts  = subtopics[0::2]
    right_sts = subtopics[1::2]

    def block_height(st):
        return max(len(st.get('assumptions', [])), 1)

    def total_height(sts):
        if not sts:
            return 0
        return sum(block_height(s) + GAP for s in sts) - GAP

    meta_height  = 1 + len(meta_assumptions) + 1
    diagram_rows = max(total_height(left_sts), total_height(right_sts), 1)
    center_row   = meta_height + diagram_rows // 2 + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diagram"

    for col_letter, w in {'A': 30, 'B': 22, 'C': 18, 'D': 22, 'E': 30}.items():
        ws.column_dimensions[col_letter].width = w
    for row_num in range(1, meta_height + diagram_rows + 2):
        ws.row_dimensions[row_num].height = 32

    # Meta-tason oletukset block
    label = ws.cell(row=1, column=1, value="Meta-tason oletukset")
    label.font = Font(bold=True, size=11, italic=True)
    label.alignment = Alignment(horizontal='left', vertical='center')
    for i, ma in enumerate(meta_assumptions):
        mc = ws.cell(row=2 + i, column=1, value=ma)
        mc.font = Font(size=10, italic=True)
        mc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Central concept
    diagram_offset = meta_height
    cc = ws.cell(row=center_row, column=3, value=central_concept)
    cc.font = Font(bold=True, size=14)
    cc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_subtopic_block(st, base_row, name_col, assump_col, assump_align):
        name_cell = ws.cell(row=base_row, column=name_col, value=st.get('name', ''))
        name_cell.font = Font(bold=True, size=11)
        name_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for j, assump in enumerate(st.get('assumptions', [])[:ROWS_PER_BLOCK]):
            ac = ws.cell(row=base_row + j, column=assump_col, value=assump)
            ac.font = Font(size=10)
            ac.alignment = Alignment(horizontal=assump_align, vertical='center', wrap_text=True)

    row = 1 + diagram_offset
    for st in left_sts:
        write_subtopic_block(st, row, name_col=2, assump_col=1, assump_align='right')
        row += block_height(st) + GAP

    row = 1 + diagram_offset
    for st in right_sts:
        write_subtopic_block(st, row, name_col=4, assump_col=5, assump_align='left')
        row += block_height(st) + GAP

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, model_used
